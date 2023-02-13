#!/usr/bin/env python
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook




def is_green(cell):

    rgb = cell.fill.start_color.rgb

    # Convierte el código RGB de hexadecimal a decimal
    red = int(rgb[2:4], 16)
    green = int(rgb[4:6], 16)
    blue = int(rgb[6:8], 16)

    # Verifica si el código RGB corresponde a alguna tonalidad de verde
    if red >= 0 and red <= 127 and green >= 128 and green <= 255 and blue >= 0 and blue <= 127:
        return True
    else:
        return False


def is_yellow(cell):

    rgb = cell.fill.start_color.rgb

    # Convierte el código RGB de hexadecimal a decimal
    red = int(rgb[2:4], 16)
    green = int(rgb[4:6], 16)
    blue = int(rgb[6:8], 16)

    # Verifica si el código RGB corresponde a alguna tonalidad de amarillo
    if red >= 128 and red <= 255 and green >= 128 and green <= 255 and blue >= 0 and blue <= 127:
        return True
    else:
        return False






class MasterChange(object):

    def __init__(self, master_file, changes_file = None):
        self.changes = None
        self.master_file = master_file
        self.changes_file = changes_file
        self.df_master = pd.read_excel(master_file, sheet_name=0)



    def get_changes(self):
        master_wb = load_workbook(self.changes_file)            # we use the function load_workbook from openpyxl
        sheet = master_wb.worksheets[0]                 # select the first sheet
        n_row, n_col = sheet.max_row, sheet.max_column  # define the rows and cols of the sheet
        headers = [c.value for c in sheet[1] if c.value is not None]
        changes = []

        for i in range(1,n_row):
            asset_changes = []
            cell = sheet.cell(column=2, row=i)
            asset_color = cell.fill.bgColor.index
            if (asset_color != '00000000'):
                effective_sheet = sheet[i][2:]              # we are interested in other columns
                for j, col in enumerate(effective_sheet):
                    col_color = col.fill.bgColor.index
                    if (col_color != '00000000'):           # if cell has color then we save the changes
                        asset_changes.append({'value':col.value,'column':headers[col.column-1]})
                changes.append({'name':sheet.cell(column=2, row=i).value ,'changes':asset_changes})

        self.changes = changes















def get_sheet_changes(file_name):
    master_wb = load_workbook(file_name)            # we use the function load_workbook from openpyxl
    sheet = master_wb.worksheets[0]                 # select the first sheet
    n_row, n_col = sheet.max_row, sheet.max_column  # define the rows and cols of the sheet
    headers = [c.value for c in sheet[1] if c.value is not None]
    changes = []

    for i in range(1,n_row):
        asset_changes = []
        cell = sheet.cell(column=2, row=i)
        asset_color = cell.fill.bgColor.index
        if (asset_color != '00000000'):
            effective_sheet = sheet[i][2:]              # we are interested in other columns
            for j, col in enumerate(effective_sheet):
                col_color = col.fill.bgColor.index
                if (col_color != '00000000'):           # if cell has color then we save the changes
                    asset_changes.append({'value':col.value,'column':headers[col.column-1]})
            changes.append({'name':sheet.cell(column=2, row=i).value ,'changes':asset_changes})

    return changes



def set_sheet_changes(changes, file_name, utc):

    df_master = pd.read_excel(file_name, sheet_name=0)

    # Here we will do the changes in the line (row)
    for change in changes:

        # if the asset is in the dataframe then we do the changes
        if change["name"] in df_master["name"].unique():

            # we define the start and end times
            t_e = datetime.utcnow().replace(hour=utc, minute=0, second=0) - timedelta(seconds=1)
            t_s = datetime.utcnow().replace(hour=utc, minute=0, second=0)

            idx = df_master.loc[df_master["name"] == change["name"]].index[0] # saving the index where there are changes

            print("id:", df_master.loc[df_master["name"] == change["name"], "id"].values[0])
            # we get the amount of void rows
            print("tamaño df:", len(df_master), "indice:", idx)
            void_rows_counter = 0
            if (idx + 1 != len(df_master)):
                while (df_master.isna().loc[idx + 1 + void_rows_counter, "name"]):
                    void_rows_counter += 1


            # iterating in all changes of certain row
            for col_change in change["changes"]:

                print("id:", int(df_master.loc[df_master["name"] == change["name"],"id"].values[0]), "asset:", change["name"], "Columna:", col_change["column"], "Valor:", col_change["value"], "n_filas:",void_rows_counter)

                # wrinting the name of columns depending on the name of the change
                ct          =   "c_" if str(col_change["column"]) == "active" else "t_"
                t_col_v     =   ct + col_change["column"] + "_v"
                t_col_ds    =   ct + col_change["column"] + "_ds"
                t_col_de    =   ct + col_change["column"] + "_de"

                void_row = pd.DataFrame([[np.nan for i in range(len(df_master.columns))]], columns=df_master.columns)


                # creating variables to do a comparison
                comparison_var_df           =       df_master.loc[df_master["name"] == change["name"], col_change["column"]].values[0]
                comparison_var_change       =       col_change["value"]
                if type(comparison_var_df)  ==  str:
                    comparison_var_change   =       str(col_change["value"]).strip()
                    comparison_var_df       =       comparison_var_df.strip()
                else:
                    comparison_var_change   =       int(comparison_var_change)
                    comparison_var_df       =       int(comparison_var_df)

                # if the previous value was different than the current one then we do the change (we compare in a string form)
                # if not, then we dont do anything
                if comparison_var_df != comparison_var_change:
                    print("deberían ser diferentes:", str(df_master.loc[df_master["name"] == change["name"], col_change["column"]].values[0]), str(col_change["value"]))

                    # here we avoid the missing values of temporal tags in rows without spaces below
                    if void_rows_counter == 0:
                        if (col_change["column"] != "active") or (df_master.loc[df_master["name"] == change["name"], col_change["column"]].values[0] == 1):
                            df_master.loc[df_master["name"] == change["name"], t_col_v] = df_master.loc[df_master["name"] == change["name"], col_change["column"]].values[0]

                    df_master.loc[df_master["name"] == change["name"], col_change["column"]] = col_change["value"]

                    # if the asset doesnt have void rows bellow then we check the line
                    if void_rows_counter == 0 :
                        print("no tengo filas iniciales")
                        # if the temporal value is different than the change then
                        print(str(df_master.loc[df_master["name"] == change["name"], t_col_v].values[0]), str(col_change["value"]))

                        # if the change is not "active"
                        if col_change["column"] != "active":
                            print("No es active")
                            # here we fill the cell
                            void_row.loc[0, t_col_v] = col_change["value"]
                            # then we concat with the new line
                            df_master = pd.concat([df_master.iloc[:idx+1+void_rows_counter], void_row, df_master.iloc[idx+1+void_rows_counter:]]).reset_index(drop=True)
                            void_rows_counter += 1
                            print("se agrega fila")


                        # if the change is "active"
                        else:
                            print("Es active")
                            # if the change is from 1 to 0
                            if col_change["value"] == 0:
                                df_master.loc[idx, t_col_de] = t_e.strftime("%Y%m%d %H:%M:%S")

                            # if the change is from 0 to 1
                            else:
                                if df_master.isna().loc[idx, t_col_v] :
                                    # filling the cell
                                    df_master.loc[idx, t_col_v] = col_change["value"]
                                    df_master.loc[idx, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")
                                else:
                                    # filling the new line
                                    void_row.loc[0, t_col_v] = col_change["value"]
                                    # then we concat with the new line
                                    df_master = pd.concat([df_master.iloc[:idx+1+void_rows_counter], void_row, df_master.iloc[idx+1+void_rows_counter:]]).reset_index(drop=True)
                                    void_rows_counter += 1
                                    # starting an active
                                    df_master.loc[idx + void_rows_counter, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")
                                    print("se agrega fila")


                    # if there are initial rows below
                    else:
                        print("si tengo filas iniciales")

                        # we check the last added row and if the cell is void we search where we have to put the value
                        if df_master.isna().loc[idx + void_rows_counter, t_col_v]:

                            # if the change is not "active"
                            if col_change["column"] != "active":
                                # we search for the first void cell
                                for i in range(void_rows_counter):

                                    if  (df_master.loc[idx + i, t_col_v] == col_change["value"]):
                                        print("Estoy vacio y el anterior si es lo mismo")
                                        break

                                    else: #(df_master.loc[idx + i, t_col_v] != col_change["value"]):
                                        df_master.loc[idx + 1 + i, t_col_v] = col_change["value"]

                                        # here we put the end and start time of changes
                                        df_master.loc[idx + i, t_col_de] = t_e.strftime("%Y%m%d %H:%M:%S")
                                        df_master.loc[idx + i + 1, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")

                                        print("Estoy vacio y el anterior no es lo mismo")
                                        print(str(df_master.loc[idx + i, t_col_v]), str(col_change["value"]))
                                        break

                            # if the change is in the feature "active"
                            else:
                                print("Es active")
                                if col_change["value"] == 1:
                                # searching for the first void cell to put the value
                                    for i in range(1, void_rows_counter+1):
                                        # if it is void then is filled
                                        if  df_master.isna().loc[idx + i, t_col_v]:
                                            # we put the value (it should be 1)
                                            df_master.loc[idx + i, t_col_v] = col_change["value"]
                                            # here we put the end and start time of changes
                                            df_master.loc[idx + i, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")
                                            break

                                else:
                                    for i in range(1, void_rows_counter + 1):
                                        # if it is void then is filled
                                        if  df_master.isna().loc[idx + i, t_col_v]:
                                            # here we put the end and start time of changes
                                            df_master.loc[idx + i - 1, t_col_de] = t_e.strftime("%Y%m%d %H:%M:%S")
                                            break


                        # if the last value is the same that the change then we dont do anything
                        elif (str(df_master.loc[idx + void_rows_counter, t_col_v]) == str(col_change["value"])) and (col_change["column"] != "active"):
                                pass


                        elif (col_change["column"] == "active"):
                            print("Es active")

                            # if the change is from 1 to 0
                            if col_change["value"] == 0:
                                df_master.loc[idx + void_rows_counter, t_col_de] = t_e.strftime("%Y%m%d %H:%M:%S")

                            # if the change is from 0 to 1
                            else:
                                # here we fill the cell
                                void_row.loc[0, t_col_v] = col_change["value"]
                                # we concat the new line
                                df_master = pd.concat([df_master.iloc[:idx+1+void_rows_counter], void_row, df_master.iloc[idx+1+void_rows_counter:]]).reset_index(drop=True)
                                void_rows_counter += 1
                                # we start the active position
                                df_master.loc[idx + void_rows_counter, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")
                                print("se agrega fila")



        else:
            # we create the new id
            new_id = int(df_master["id"].max() + 1)
            # creating the the new row
            void_row = pd.DataFrame([[np.nan for i in range(len(df_master.columns))]], columns=df_master.columns)
            # setting the id
            void_row.loc[0, "id"]   =   new_id
            void_row.loc[0, "name"] =   change["name"]
            # passing for the changes
            for col_change in change["changes"]:
                ct          =   "c_" if str(col_change["column"]) == "active" else "t_"
                t_col_v     =   ct + col_change["column"] + "_v"
                void_row.loc[0, col_change["column"]] = col_change["value"]

                # we check if the atribute has a temporal tag
                if t_col_v in df_master.columns:
                    void_row.loc[0, t_col_v] = col_change["value"]
            # adding the new line
            df_master = pd.concat([df_master, void_row]).reset_index(drop=True)

            print("se agrega asset,", "id:", new_id, "asset:", change["name"])

























# ESTO TIENE QUE SER UN MÉTODO DE CLASE

# ------ WRITE AND SAVE THE NEW FILE ------ #

    # it takes the file
    book = load_workbook(file_name)

    # creating a date string
    str_date = datetime.now().strftime("%Y%m%d")
    # naming the sheet
    new_sheet_name = "Maestro " + str_date
    # naming the file
    new_master_name = "maestro_" + str_date + "_NEW.xlsx"

    # we create a new work book
    writer = pd.ExcelWriter(new_master_name, engine='openpyxl')

    # we write on it
    writer.book = book

    # seting the new dataframe in the excel file
    df_master.to_excel(writer, index=False, sheet_name=new_sheet_name)

    sheet_names_ = writer.book.sheetnames

    # here we are putting the new sheet at the beginning of the excel file
    writer.book._sheets = [writer.book[new_sheet_name]] + [writer.book[name] for name in sheet_names_ if name != new_sheet_name]
    #_len = len(sheet_names_)
    # sheet_names_[-1]
    # [writer.book[sheet_names_[i]] for i in range(0, _len-1)]

    # Finaly we save the dataframe
    writer.close()

