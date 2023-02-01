#!/usr/bin/env python
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook




def get_sheet_changes(file_name):
    master_wb = load_workbook(file_name)            # we use the function load_workbook from openpyxl
    sheet = master_wb.worksheets[0]                 # select the first sheet
    n_row, n_col = sheet.max_row, sheet.max_column  # define the rows and cols of the sheet
    headers = [c.value for c in sheet[1] if c.value is not None]
    changes = []

    for i in range(1,n_row):
        asset_changes = []
        cell = sheet.cell(column=2, row=i)
        asset_color = cell.fill.bgColor.index
        if (asset_color != '00000000'):
            effective_sheet = sheet[i][2:]              # we are interested in other columns
            for j, col in enumerate(effective_sheet):
                col_color = col.fill.bgColor.index
                if (col_color != '00000000'):           # if cell has color then we save the changes
                    asset_changes.append({'value':col.value,'column':headers[col.column-1]})
            changes.append({'name':sheet.cell(column=2, row=i).value ,'changes':asset_changes})

    return changes






def set_sheet_changes(changes, file_name, utc):

    df_master = pd.read_excel(file_name, sheet_name=0)

    # Here we will do the changes in the line (row)
    for change in changes:

        t_e = datetime.utcnow().replace(hour=utc, minute=0, second=0) - timedelta(seconds=1)
        t_s = datetime.utcnow().replace(hour=utc, minute=0, second=0)

        idx = df_master.loc[df_master["name"] == change["name"]].index[0] # saving the index where there are changes


        # we get the amount of void rows
        void_rows_counter = 0
        while(df_master.isna().loc[idx + 1 + void_rows_counter, "name"]):
            void_rows_counter += 1



        for col_change in change["changes"]:    # iterating in all changes of certain row

            print("id:", int(df_master.loc[df_master["name"] == change["name"],"id"].values[0]), "asset:", change["name"], "Columna:", col_change["column"], "Valor:", col_change["value"], "n_filas:",void_rows_counter)

            ct          =   "c_" if str(col_change["column"]) == "active" else "t_"
            t_col_v     =   ct + col_change["column"] + "_v"
            t_col_ds    =   ct + col_change["column"] + "_ds"
            t_col_de    =   ct + col_change["column"] + "_de"

            # If the previous value was different than the current one then we do the change (we compare in a string form)
            if (str(df_master.loc[df_master["name"] == change["name"], col_change["column"]].values[0]) != str(col_change["value"])):

                # here we avoid the missing values of temporal tags in rows without spaces below
                if void_rows_counter == 0:
                    df_master.loc[df_master["name"] == change["name"], t_col_v] = df_master.loc[df_master["name"] == change["name"], col_change["column"]].values[0]
                df_master.loc[df_master["name"] == change["name"], col_change["column"]] = col_change["value"]

                # if the asset doesnt have void rows bellow then we check the line
                if void_rows_counter == 0 :
                    print("no tengo filas iniciales")

                    if (df_master.loc[df_master["name"] == change["name"], t_col_v].values[0] != col_change["value"]):
                        print(str(df_master.loc[df_master["name"] == change["name"], t_col_v].values[0]), str(col_change["value"]))
                        void_row = pd.DataFrame([[np.nan for i in range(len(df_master.columns))]], columns=df_master.columns)
                        void_row.loc[0, t_col_v] = col_change["value"] # here we fill the cell

                        df_master = pd.concat([df_master.iloc[:idx+1+void_rows_counter], void_row, df_master.iloc[idx+1+void_rows_counter:]]).reset_index(drop=True)
                        void_rows_counter += 1

                        df_master.loc[idx + void_rows_counter - 1, t_col_de] = t_e.strftime("%Y%m%d %H:%M:%S")
                        df_master.loc[idx + void_rows_counter, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")

                        print("se agrega fila")



                # if there are initial rows below
                else :
                    print("si tengo filas iniciales")
                    print("valor antes:", df_master.loc[idx + void_rows_counter, t_col_v], "valor ahora:", col_change["value"])

                    # we check the last added row and if the cell is void we search where we have to put the value
                    if df_master.isna().loc[idx + void_rows_counter, t_col_v]:
                        for i in range(void_rows_counter):

                            if  (df_master.loc[idx + i, t_col_v] == col_change["value"]):
                                print("Estoy vacio y el anterior si es lo mismo")
                                break

                            else: #(df_master.loc[idx + i, t_col_v] != col_change["value"]):
                                df_master.loc[idx + 1 + i, t_col_v] = col_change["value"]

                                # here we put the end and start time of changes
                                df_master.loc[idx + i, t_col_de] = t_e.strftime("%Y%m%d %H:%M:%S")
                                df_master.loc[idx + i + 1, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")

                                print("Estoy vacio y el anterior no es lo mismo")
                                print(str(df_master.loc[idx + i, t_col_v]), str(col_change["value"]))
                                break


                    # if the last value is the same that the change then we dont do anything
                    elif str(df_master.loc[idx + void_rows_counter, t_col_v]) == str(col_change["value"]) :
                        pass

                    # if the last cell is filled then we add a void row and we put the value
                    else:
                        void_row = pd.DataFrame([[np.nan for i in range(len(df_master.columns))]], columns=df_master.columns)
                        void_row.loc[0, t_col_v] = col_change["value"] # here we fill the cell

                        df_master = pd.concat([df_master.iloc[:idx+void_rows_counter+1], void_row, df_master.iloc[idx+void_rows_counter+1:]]).reset_index(drop=True)
                        #print(df_master.iloc[:idx+void_rows_counter, t_col_v].tail(1))
                        void_rows_counter += 1
                        # here we put the end and start time of changes
                        df_master.loc[idx + void_rows_counter - 1, t_col_de] = t_e.strftime("%Y%m%d %H:%M:%S")
                        df_master.loc[idx + void_rows_counter, t_col_ds] = t_s.strftime("%Y%m%d %H:%M:%S")

                        print("se agrega fila")

# ------ WRITE AND SAVE THE NEW FILE ------ #

    # it takes the file
    book = load_workbook(file_name)

    # creating a date string
    str_date = datetime.now().strftime("%Y%m%d")
    # naming the sheet
    new_sheet_name = "Maestro " + str_date
    # naming the file
    new_master_name = "maestro_" + str_date + "_NEW.xlsx"

    # we create a new work book
    writer = pd.ExcelWriter(new_master_name, engine='openpyxl')

    # we write on it
    writer.book = book

    # seting the new dataframe in the excel file
    df_master.to_excel(writer, index=False, sheet_name=new_sheet_name)

    sheet_names_ = writer.book.sheetnames

    # here we are putting the new sheet at the beginning of the excel file
    _len = len(sheet_names_)
    writer.book._sheets = [writer.book[sheet_names_[-1]]] + [writer.book[sheet_names_[i]] for i in range(0, _len-1)]

    # Finaly we save the dataframe
    writer.close()

